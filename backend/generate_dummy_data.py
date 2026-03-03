#!/usr/bin/env python3
"""
Standalone script to generate dummy commodity price data and save as CSV, XLSX, or JSON.
Output is saved under backend/generated_data/. Data can include sharp price moves so that
when ingested, the platform will create alerts/risks (e.g. price spike/drop >10%).

Not part of the application — run directly: python generate_dummy_data.py

Requirements:
  - Python 3.7+
  - For XLSX output: pip install openpyxl

Usage:
  python generate_dummy_data.py
  Then follow the prompt to choose format (csv, xlsx, json), optional filename,
  and whether to include alert/risk triggers (sharp price moves).
"""

import csv
import json
import random
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List

# Commodities with base price (₹) and daily volatility for realistic variation
COMMODITIES = [
    ("Wheat", 2100, 0.02),
    ("Rice", 3200, 0.015),
    ("Cabbage", 280, 0.04),
    ("Brinjal", 850, 0.03),
    ("Potato", 450, 0.025),
    ("Cucumber", 3200, 0.035),
    ("Green Chilli", 14000, 0.04),
    ("Banana", 600, 0.02),
    ("Tomato", 2200, 0.05),
    ("Onion", 1800, 0.03),
]
REGIONS = ["Mumbai", "Delhi", "Bangalore", "Chennai", "Kolkata"]
DEFAULT_DAYS = 90
# Threshold used by dashboard to create alerts (e.g. 10% move)
ALERT_CHANGE_PCT = 10.0
# All generated files go here (same folder as generate_alerts_risks_data.py output)
OUTPUT_DIR = Path(__file__).resolve().parent / "generated_data"


def _apply_alert_triggers(
    commodity_rows: List[Dict[str, Any]],
    num_spikes: int = 2,
    num_drops: int = 2,
) -> None:
    """
    Mutate commodity_rows in place: inject sharp one-day moves so that when this data
    is ingested, dashboard/ingestion logic will create price_spike / risk alerts.
    """
    if len(commodity_rows) < 2:
        return
    # Pick random day indices (avoid first day so we have a "previous" price)
    indices = list(range(1, len(commodity_rows)))
    random.shuffle(indices)
    # Spike: e.g. +12% to +22%
    for i in indices[:num_spikes]:
        prev_price = float(commodity_rows[i - 1]["price"])
        pct = random.uniform(ALERT_CHANGE_PCT, 22) / 100.0
        commodity_rows[i]["price"] = round(prev_price * (1 + pct), 2)
    # Drop: e.g. -12% to -20%
    for i in indices[num_spikes : num_spikes + num_drops]:
        prev_price = float(commodity_rows[i - 1]["price"])
        pct = random.uniform(-0.20, -ALERT_CHANGE_PCT / 100.0)
        commodity_rows[i]["price"] = round(max(prev_price * (1 + pct), 10), 2)


def generate_rows(
    num_days: int = DEFAULT_DAYS,
    include_region: bool = True,
    include_volume: bool = True,
    include_alert_triggers: bool = True,
) -> List[Dict[str, Any]]:
    """Generate list of dicts: commodity, price, date, [region], [volume].
    If include_alert_triggers is True, some rows get sharp price moves so ingestion creates alerts/risks.
    """
    end_date = datetime.utcnow().date()
    rows = []
    for name, base_price, volatility in COMMODITIES:
        price = float(base_price)
        commodity_rows = []
        for d in range(num_days):
            record_date = end_date - timedelta(days=d)
            price = price * (1 + random.gauss(0, volatility))
            price = round(max(price, 10), 2)
            row = {
                "commodity": name,
                "price": price,
                "date": record_date.strftime("%Y-%m-%d"),
            }
            if include_region:
                row["region"] = random.choice(REGIONS)
            if include_volume:
                row["volume"] = round(random.uniform(50, 500), 2)
            commodity_rows.append(row)
        if include_alert_triggers and len(commodity_rows) >= 2:
            _apply_alert_triggers(commodity_rows, num_spikes=2, num_drops=2)
        rows.extend(commodity_rows)
    return rows


def save_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)


def save_json(rows: List[Dict[str, Any]], path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)


def save_xlsx(rows: List[Dict[str, Any]], path: Path) -> None:
    try:
        import openpyxl
    except ImportError:
        print("XLSX support requires openpyxl. Install with: pip install openpyxl")
        return
    if not rows:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price History"
    headers = list(rows[0].keys())
    for col, key in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=key)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            val = row[key]
            if isinstance(val, (int, float)) and key == "price":
                ws.cell(row=row_idx, column=col_idx, value=val)
            else:
                ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(path)


def main():
    print("Dummy commodity price data generator")
    print("------------------------------------")
    num_days = input(f"Number of days of data per commodity [{DEFAULT_DAYS}]: ").strip() or str(DEFAULT_DAYS)
    try:
        num_days = int(num_days)
        num_days = max(1, min(num_days, 365 * 2))
    except ValueError:
        num_days = DEFAULT_DAYS

    include_region = input("Include 'region' column? [Y/n]: ").strip().lower() != "n"
    include_volume = input("Include 'volume' column? [Y/n]: ").strip().lower() != "n"
    include_alert_triggers = (
        input("Include alert/risk triggers (sharp price moves so ingestion creates alerts)? [Y/n]: ").strip().lower() != "n"
    )

    print("\nOutput format: 1=CSV, 2=XLSX, 3=JSON")
    choice = input("Choose format (1/2/3) [1]: ").strip() or "1"

    default_name = "dummy_price_data"
    name = input(f"Output filename (without extension) [{default_name}]: ").strip() or default_name
    name = name.replace(" ", "_")

    rows = generate_rows(
        num_days=num_days,
        include_region=include_region,
        include_volume=include_volume,
        include_alert_triggers=include_alert_triggers,
    )
    print(f"Generated {len(rows)} rows for {len(COMMODITIES)} commodities.", end="")
    if include_alert_triggers:
        print(" Includes sharp price moves (alert/risk triggers).")
    else:
        print()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if choice == "2":
        ext = "xlsx"
        path = OUTPUT_DIR / f"{name}.{ext}"
        save_xlsx(rows, path)
    elif choice == "3":
        ext = "json"
        path = OUTPUT_DIR / f"{name}.{ext}"
        save_json(rows, path)
    else:
        ext = "csv"
        path = OUTPUT_DIR / f"{name}.{ext}"
        save_csv(rows, path)

    if path.exists():
        print(f"Saved: {path}")
    else:
        print("Save failed or skipped (e.g. missing openpyxl for XLSX).")


if __name__ == "__main__":
    main()
